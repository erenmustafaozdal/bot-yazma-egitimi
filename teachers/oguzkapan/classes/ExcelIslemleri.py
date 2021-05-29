## Deneysel çalışma
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
class ExcelIslemleri:
    def wb_ac(xl_path):
            if not os.path.exists(xl_path):
                wb = Workbook()
            else:
                wb = load_workbook(xl_path)
            return wb
