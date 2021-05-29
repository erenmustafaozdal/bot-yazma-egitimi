from selenium import webdriver
import settings
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from openpyxl import Workbook, load_workbook
from teachers.oguzkapan.classes.browser import Browser
from teachers.oguzkapan.classes.eba import EBA
import os

## Amaç : Tc kimlik numaraları bir excel dosyasında kayıtlı olan öğrencilere tek kullanımlık şifre verme.
## Excel Sütünları aşağıdaki gibi düşünülmüştür.
## Okul No / Ad Soyad / Tc Kimlik No / Şifre

## Eba'ya giriş yap
browser = Browser(settings.driver_path)
driver = browser.get()

eba = EBA(driver)
eba.login(settings.tc, settings.password)


def profil_sayfasina_gec():
    # Profil menüsünü aç
    driver.find_element_by_xpath("//div[@id='vcProfileWidget']//div[@class='vc-profile-widget-caret vc-position-relative']").click()
    time.sleep(5)
    # Şifre Bölümüne Giriş - Öğretmen
    driver.find_element_by_xpath("//*[@id='eba-menu-panel']/div/div[3]/a/div").click()
    time.sleep(3)

profil_sayfasina_gec()
i=2 ##  Excel dosyasında ilk satırda sütün başlıkları olduğu için döngü 2. satırdan başlatıldı.
## Eba'ya giriş yapıldı

## Excel döküman ayarlarını yap.
excel_dir="./excels/eba_sifre_ver/"
xl_path = f"{excel_dir}/eba_ogrenci_bilgileri.xlsx"
if not os.path.exists(xl_path):
    wb = Workbook()
else:
    wb = load_workbook(xl_path)
ws=wb.active
kontrol=0 ## Uyarıda yer alan bir daha gösterme onay kutusuna tıklama işlemini kontrol eder.
#excel dökümanında kayıtlı öğrencilerin tc_kimlik_numaralarını al şifre üret ve karşısına yaz
while i<ws.max_row+1:
    ogrenci_tc_kimlik_no=ws.cell(i, 3).value

    # Bir daha gösterme onay kutusunu tıkla
    if(kontrol==0):
        bir_daha_gosterme=driver.find_element_by_xpath("// input[ @ name = 'modal-gun']").click()
        time.sleep(2)
        kontrol=1
        uyarıyı_kapat = driver.find_element_by_xpath('//*[@id="uyari"]/div/div/div[1]/button').click()
        time.sleep(1)

    inputbox = driver.find_element(By.ID, 'studentTckn').send_keys(ogrenci_tc_kimlik_no)
    bilgigetir = driver.find_element_by_xpath('//*[@id="get-tckn"]/button').click()
    time.sleep(2)

    #Şifre Oluştur
    tek_sifre = driver.find_element_by_xpath("//button[contains(text(),'Tek Kullanımlık Giriş Şifresi Oluştur')]").click()
    time.sleep(1)

    yeni_sifre = driver.find_element_by_xpath("//code[@id='user-pass']").text
    time.sleep(2)

    ##
    driver.find_element_by_xpath('// *[ @ id = "div-user-pass"] / a').click()

    time.sleep(1)
    ws[f"D{i}"]=yeni_sifre ## Excel aktif çalışma kitabının D sütünu i. satırına yaz. (i Excel tablosundaki max_row özelliğine kadar artacak şekilde ayarlandı)
    i+=1

wb.save(xl_path)
driver.close()


"""
Deneysel Çalışma 
İstenen Sınıftaki öğrencilerin Tc numaralarını otomatik al ve Excel'e yaz 

eokul giriş yapıldıktan sonra 
    ortaogretim_kurumlari_giris= driver.find_element_by_xpath("//span[contains(text(),'Ortaöğretim Kurum İşlemleri')]").click()
    not_islemleri=driver.find_element_by_xpath("//span[contains(text(),'Not İşlemleri')]").click()
    ders_notu_girisi=driver.find_element_by_xpath("//a[contains(text(),'Ders Notu Girişi')]").click()
    sinif_sec=driver.find_element_by_xpath("//option[@value='7056750']").click()
    
    ders_sec_select_box=//select[@id='ddlDersler']
    herhangi_bir_dersi_al=//option[@value='1100'] ## burada temel mantık listedeki herhangi bir dersi tıklaması 
    listeleye_tikla=//input[@id='btnListele']
    
def table_is_loaded():
    try:
       return wait.until(
            EC.visibility_of_all_elements_located((By.XPATH, "//div[@class='body-container']/div[@role='row']"))
            )
    except:
        print("Tablo yüklenemedi.")
        driver.refresh()
        return  table_is_loaded()
        
        
        
    students=table_is_loaded()
    student_count=len(students)
    
    tabloda öğrenci noları sütunu:
    tabloda öğrenci adları sütunu : //body[1]/form[2]/table[1]/tbody[1]/tr[3]/td[2]/table[1]/tbody[1]/tr[3]/td[1]/table[1]/tbody[1]/tr[9]/td[2]/table[1]/tbody[1]/tr[1]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/table[1]/tbody[1]/tr[2]/td[1]/table[1]/tbody[1]/tr[2]/td[2]
    

"""