"""
StudentBasedReportV3Openxyl
Oguz Kapan
"""
import datetime

from selenium import webdriver
import settings
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime


def left_menu_is_loaded():
    try:
        wait.until(EC.visibility_of_all_elements_located((By.ID,"vc-treeleftmenu")))
    except:
        print("Menu yüklenirken çok bekledi. Sayfa yenileniyor...")
        driver.refresh()
        left_menu_is_loaded()

def login(tc,password):
    driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")
    if 'VCollabPlayer' in driver.current_url:
        return
    driver.find_element_by_xpath("//button[@title='edevlet girişi']").click()
    driver.find_element_by_css_selector("#tridField").send_keys(settings.tc)
    driver.find_element_by_id("egpField").send_keys(settings.password)
    try:
        driver.find_element_by_xpath("//input[@name='submitButton']").click()
    except:
      print("Bazı hatalar:D::D - Özetle giriş yapılamadı")
      login(tc,password)
    else:
        left_menu_is_loaded()


# tablo satırlarını al
def table_is_loaded():
    try:
       return wait.until(
            EC.visibility_of_all_elements_located((By.XPATH, "//div[@class='body-container']/div[@role='row']"))
            )
    except:
        print("Tablo yüklenemedi.")
        driver.refresh()
        return  table_is_loaded()

img_dir="./images/student-based-reports"
if not os.path.exists(img_dir):
    os.mkdir(img_dir)

excel_dir="./excels/student-based-reports"
if not os.path.exists(excel_dir):
    os.mkdir(excel_dir)

def excel_wb(xl_path):
    if not os.path.exists(xl_path):
        wb = Workbook()
    else:
        wb = load_workbook(xl_path)
    return wb

xl_path = f"{excel_dir}/student-based-reports_nn.xlsx"
wb=excel_wb(xl_path)

##dersler=["Web Tasarım ve Programlama","Grafik ve Animasyon","Açık Kaynak İşletim Sistemi"]
dersler=["Web Tasarım ve Programlama"]
def excel_sheet_setup(ders_adi):
    ws = wb.create_sheet(ders_adi)
    ws.append(["Tarih", "Öğrenci", "Tamamlama", "Performans", "Ekran Görüntüsü"])
    ##ws.set_column(0, 10, 25)

    return ws

def excele_veri_yolla(ws,satir,sutun,veri):
   ws.write(satir,sutun,veri)

sutun=0
satir=1

## **********Tarayıcı nesnesi********************* ##
driver=webdriver.Chrome(settings.driver_path)
driver.maximize_window()
driver.set_page_load_timeout(10)
## **********Tarayıcı nesnesi********************* ##

# poll_frequbcey ne kadarda bir deneme yapılaca k##
wait = WebDriverWait(driver, timeout=10, poll_frequency=1)
## poll_frequbcey ne kadarda bir deneme yapılacak ##

##** Eba giriş **##
login(settings.tc,settings.password)
##** Eba giriş **##

##** Raporlar **##
reports_menu=wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@class,'vc-lm-item-title')][normalize-space()='Raporlar']")))
reports_menu.click()
##** Raporlar **##

##** Çalışma Raporları **##
work_reports=wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()='Çalışma Raporları']")))
work_reports.click()
##** Çalışma Raporları **##

##** Çalışma Raporları--> Gelen tabloyu al **##
table_is_loaded()
##** Çalışma Raporları--> Gelen tabloyu al **##

##** Öğrenci Bazlı Kısmına Geç **##
student_based_reports_take=wait.until(EC.element_to_be_clickable((By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")))
student_based_reports_take.click()
time.sleep(2)
##** Öğrenci Bazlı Kısmına Geç **##


""" 
listbox içerisindeki verilere ulaşmak için kullanılabilecek xpath bilgileri
//option[@label='Web Tasarım ve Programlama']
//option[text()='11. Sınıf']
//option[contains(text(),'11. Sınıf')]
"""

##** Sınıf seçimini yaptır. **##
class_select=wait.until(EC.element_to_be_clickable((By.XPATH, "//option[contains(text(),'11. Sınıf')]")))
class_select.click()

##** Sınıf seçimini yaptır. **##

"""
#Tanımlama yapılan tek sınıf olduğu için  ihtiyaç olmadığı tespit edildi.
# driver.find_element_by_xpath("//option[@label='AMP - 11. Sınıf / A Şubesi (BİLİŞİM TEKNOLOJİLERİ ALANI)']").click()
##
# DERSLERİ LİSTEDEN ALMADAN EVVEL KULLANDIĞIMIZ YÖNTEM
# driver.find_element_by_xpath("//option[@label='Grafik ve Animasyon']").click()
"""  ## ihtiyaç olmayan alanlar

index=0 ## Ders sayısının takibi için oluşturulan değişken

for index in range(len(dersler)):
    satir=1 # Her dersin bilgisi alındıktan sonra tablo satır sayısını sıfırlamak için kullanılır.
    lesson_select=wait.until(EC.element_to_be_clickable((By.XPATH, f"//option[@label='{dersler[index]}']")))
    lesson_select.click()
    ##driver.find_element_by_xpath(f"//option[@label='{dersler[index]}']").click()
   # time.sleep(2)



    ws=excel_sheet_setup(dersler[index])
    students=table_is_loaded()
    student_count=len(students)

    # Günün tarihini al
    date = datetime.today()
    # Excel'de yazılmış en son satırı al

    last_row = ws.max_row
    print(last_row)
    for student_i in range(student_count):
            students=table_is_loaded()
            students[student_i].click()

            #çalışma satırlarını al
            works=table_is_loaded()
            student_name=driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text

            completes=[]
            performances=[]

            for work_i in range(4):
               complete= works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
               completes.append(int(complete.replace("%","")))
            performance=works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
            if performance!='-':
                performances.append(int(performance))
            complete_avg=sum(completes) / len(completes)
            performance_avg="performans yok"
            if len(performances)>0:
                performance_avg=sum(performances)/len(performances)

            img_path=f"{img_dir}/{student_name}.png"
            driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)


            row = last_row + student_i + 1
            ws[f"A{row}"] = date
            ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
            ws[f"B{row}"] = student_name
            ws[f"C{row}"] = complete_avg
            ws[f"D{row}"] = performance_avg
            # =KÖPRÜ("../images/student-based-reports/Ekran Görüntüsü.png";"Görüntü")
            ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'





            satir+=1
            driver.back()




# //div[contains(@class,'vc-lm-item-title')][normalize-space()='Raporlar']

# div[id='3e4fc4a8-9d2f-3518-e3d2-5cf3f3bdb571'] div[class='vc-lm-item-title ']

# driver.find_element_by_xpath("//div[contains(@class,'vc-lm-item-title')][normalize-space()='Raporlar']").click()

time.sleep(2)
driver.close()

wb.save(xl_path)
wb.close()