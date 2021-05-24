"""
Performans Ortalaması Hesap Hatası Çözümü
openpyxl ve datetime Modüllerini Tanıtma
Excel dosyasını oluşturma veya varolan dosyayı yükleme
Öğrenci çalışma bilgilerini Excel'e yazma
Excel'e kayıt sonrası çalışmanın incelenmesi
Çalışma sonrası Excel dosyasını inceleme - Excel'de neler yapılabilir
3. Excell'e verileri yazdırma
"""


from selenium import webdriver
import settings
import  time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import os # işletim sistemi çin modülü import ediyoruz
# Öncelikle openpyxl paketi yüklenir: pip install openpyxl
# Excell nesnesi için metot ekliyoruz
from openpyxl import Workbook
# Excell nesnemiz olduktan sonra bunu ekleyecek modülü ekliyoruz
from  openpyxl import load_workbook
# Günün tarihini eklemek için modül ekliyoruz
from datetime import  datetime

# SORUN ÇÖZÜMÜ: "EBA yükleniyor" mesajının gitmemesi (her sayfada karşılaşılabiliyor)
# ÇÖZÜM: Sol menü yüklenene kadar bekle
# Bunu için sol menünün yüklenip yüklenmediğini kontrol edecek bir fonksiyon tanımlıyoruz
def LeftMenuIsLoaded():
    try:
        # ID'si vc-treeleftmenu olan nesne yüklenene kadar bekle
        wait.until(ec.visibility_of_element_located((By.ID,"vc-treeleftmenu")))
    except: # 10 saniye bekledikten sonra bir hata alırsa
        print("Menünün yüklenmesi için çok bekledi. Sayfa yenileniyor...")# yazısını yazacak
        driver.refresh() # sayfayı yeniden yükleyecek
        LeftMenuIsLoaded() #yenilenebilir bir fonksiyon olarak kendini tekrar çağıracak
    # EBA yükleninceye kadar sonsuz döngüye sokuyoruz

# EBA'ya giriş işlemini gerçekleştiren fonksiyonu yazıyoruz
def login(tc, password):
    # Öğretmen giriş sayfasına gidiyoruz
    driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")

    # Eğer daha önce giriş yapıldı ise fonksiyondan çık
    # Daha öncede giriş EBA'ya giriş yapılmış ama halen aynı profilden EBA'dan çıkılmamışsa
    # EBA aynı profile tekrar yollar. Burada tekrar giriş yapma işlemini apmamak için bunu sorguluyoruz
    # 'VCollabPlayer' ifadesi EBA'ya giriş yaptıktan sonra her URL^de ortak bir değerdir,
    #Geçerli URL içinde varlığı sorgulanarak, giriş yapılıp yapılmadığı kontrol edilir
    if 'VCollabPlayer' in driver.current_url: # current_url'de VCOllabPlayer varsa fonksiyondan çık
        return

    # 'edevlet girişi' başlıklı tuşa basıyoruz ve E-Devlet giriş sayfasına gidiyoruz
    driver.find_element_by_xpath("//button[@title='edevlet girişi']").click()

    # TC nosunu "#tridField" ccs selectörü belilenen girişe yazıdıryoruz
    driver.find_element_by_css_selector("#tridField").send_keys(settings.tc)
    # E-devlet Şifresini "gpField" id ile belirlenen girişe yazıdıryoruz
    driver.find_element_by_id("egpField").send_keys(settings.password)

    # E-devlete giriş formunu gönderiyoruz
    # Eğer sayfa yüklenemez veya başka bir hata alınırsa giriş işlemini tekrar ettiriyoruz
    try:
        # e-devlet butonu dallanma elmanını selector ile tespit edip e-devlet sayfasına gidiyoruz
        driver.find_element_by_xpath("//input[@name='submitButton']").click()
    except:
        print("Sayfa 10 saniyede yüklenmedi.Sayfa yenileniyor...")
        # Sayfa yüklenmemişse fonksiyon kendini tekrar çağırıyor
        login(tc,password)
    # Eğer try bölümünde bir hata almazsa buraya geçiyor
    else:
        # Eğer başarılı bir şekilde EBA'ya giriş
        #yapıldı ise sol menünün yüklenmesini bekle
        LeftMenuIsLoaded()

#Tablo satırlarının görünmesini bekleyen fonksiyon
def TableIsLoaded():
    try: # Bütün elemanlar görünür oluncaya kadar bekle ve döndür
        return wait.until(ec.visibility_of_all_elements_located(
            (By.XPATH,"//div[@class= 'body-container']/div[@role='row']")
        ))
    except: # hata varsa
        print("Tablo yüklenemedi. Sayfa yenileniyor...")
        driver.refresh()
        # Öğrenci bazlı adım adım rapor alınırken sonsuz döngüyü engelemek için return eklendi
        return TableIsLoaded() # Sayfa yüklemedi ise özyinele.


# Ekran örüntüleri için yoksa kalsörü oluştur
# "./images/student-based-reports"
# klasör için değişken
imgDir="./images/student-based-reports"
# ./images/student-based-reports kalsörü yoksa
if not os.path.exists(imgDir):
    # kalsörü oluştur
    os.mkdir(imgDir)


# Excel dosyası eğer yoksa oluştur
# Dosya yolu için değişken
xlPath = "./excels/student-based_reports.xlsx"

# Eğer dosya yoksa
if not os.path.exists(xlPath):
    wb = Workbook() #Çalışma kitabı oluştur
    ws = wb.active #Açık olan çalışma kitabında ki çalışma sayfasının adını sakla
    ws.title="Öğrenci bazlı çalışma raporu"#Çalışma sayfasının adını değiştir
    #1. satırın başlıklarını değiştiriyoruz
    ws .append(["Tarih", "Öğrenci", "Tamamlama", "Performans", "Ekran Görüntüsü"])
else: # değilse
    # aktif olan workbook'u alıyoruz
    wb = load_workbook(xlPath)
    # "Öğrenci Bazlı Çalışma Raporları" isimli çalışma safasını alıyoruz
    ws = wb["Öğrenci Bazlı Çalışma Raporları"]

# tarayıcı nesnesini oluşturuyoruz ve değişkene atıyoruz
driver = webdriver.Chrome(settings.driver_path)
# sayfayı maksimize yapıyoruz
driver.maximize_window()
#sayfanın yüklenmesini çok beklememek için
#20 saniye beklemesini yoksa hata vermesini elirliyoruz
#Sayfanı yüklenmesini bekletiyoruz
driver.set_page_load_timeout(20)

#Çeşitli sayfa elemanlarını beklemek için;
# WebDrverWait sınıfından driver sayfasında 10 sn bekleyen 1 sık aralıkla kontrol yapan wait nesnesi üretiyoruz
# Elemanların yüklenmesini bekletiyoruz
wait = WebDriverWait(driver, timeout=10, poll_frequency=1)

# EBA'ya giriş yapılır
login(settings.tc,settings.password)

# sol menüden "Raporlar menüsü
reportsMenu=wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
))

reportsMenu.click()

#'Raporlar' sayfasında 'Çalışma Raporları' bağlantısına tıklatıyoruz.
# Çalışma Raporları butonu tıklanabilir oluncaya kadar beletiyoruz
workReports = wait.until(ec.element_to_be_clickable(
    (By.XPATH,"//div[text()='Çalışma Raporları']")
))
workReports.click()

# Bu sayfanın yüklendiğinden emin olmak için tablo satırları yüklenene kadar bekle
# TableIsLoaded() sayfası kişisel sayfamızda Çalışma Tabloları olmadığı için pasif duruma getiriyorum.:) KB
#
#TableIsLoaded() # foksiyonu çağırıyoruz
#
# 'Çalışma Raporları' sayfasında 'ÖĞRENCİ BAZLI' bağlantısına tıkla.

studentBasedLink = wait.until(ec.element_to_be_clickable(
    (By.XPATH,"//div[text()='ÖĞRENCİ BAZLI']")
))
studentBasedLink.click()

# Bu sayfanın yüklendiğinden emin olmak için tablo satırları yüklenene kadar bekle.
# Öğrencilerin olduğu tablo satırlarını al
students= TableIsLoaded() # gelen tablo bilgilerini students değişkeninde saklıyoruz

# TODO: Öğrencilerin satırlarını döngüye al.
# 1. Öğrencinin adını al.
# 2. Son 10 çalışmanın tamamlanma yüzdelerini bir diziye aktar.
# 3. Son 10 çalışmanın performanslarını bir diziye aktar.
# 4. örtlimalfli tamamlamayı hesap et.
# 5. Ortalama performansı hesap et.
# 6. Ekran görüntüsünü kaydet.
# 7. Verileri ekrana yazdır.
# 8. Verileri Excel dosyasına yazdır.

# Öğrenci sayısını bulup bunu studentCount adında ki değişkende saklıyoruz
studentCount=len(students)

# Günün talihini al
date = datetime.today()
#	yazılmış, en son satırı al
lastRow=ws.max_row

# Öğrencileri, öğrenci sayısı tek tek gezip tablonun yüklenmesini bekliyoruz
for studentI in  range(studentCount):
    students=TableIsLoaded()
    students[studentI].click()

    # çalışma satırlarını al
    works = TableIsLoaded()

    # Tabloda ki o anki öğrenci adını sudentName değişkeninde saklıyoruz
    studentName= driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text

    # Çalışma satırlarını gez ve hesaplamaları ap
    completes=[] # Tamamlama yüzdesinide ki her elemanı saklayacağımız dizi
    performances=[] # Performans notlarını saklayacağımız dizi

    # Öğrencinin ilk 10 çalışmasını tarıyoruz
    for workI in range(10):
        # O an ki çalışmanın tamamlama yüzdesini complete değişkeninde saklıyoruz
        complete = works[workI].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
        # O an ki taradığımız notunu performance değişkeninde saklayoruz
        performance = works[workI].find_element_by_css_selector("div#multiColouredProgress span").text

        # completes dizisine atarken complete değişkeninde okuduğumuz % kaldırıp integer'a çeviriyoruz
        completes.append(int(complete.replace("%","")))
        # Eğer performans notunda - farklı bir not varsa performances dizsine ekliyoruz
        if performance !="-":
            performances.append(int(performance))

    # tamamlama ortalamasını alıyoruz
    completeAvg = sum(completes) / len(completes)
    # performanş ortalamasını alıyoruz
    # performans değeri hiç olmadığında varsayım değer
    performanceAvg = "performans yok"
    # performans değeri 9'dan büyükse yani not girilmişse performans ortalamasını hesaplıyoruz
    if len(performances) > 0:
        performance_avg = sum(performances) / len(performances)

    # ekran görüntüsü alalım - Yeniden düzenlendi
    # Ekran Görüntüsü Adı Formatı: 20210508 - Öğrenci Adı.png
    imgPath = f"{imgDir}/{date.strftime('%Y%m%d')}-{studentName}.png"
    # Ekran görüntüsünü alıyoruz
    driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(imgPath)

    print("*" * 50)
    print("Öğrenci: ", studentName)
    print("--- Tamamlama:", completeAvg)
    print("— Performans:", performanceAvg)

    # ExçeVe yazdır
    #(last_row=1) + (student_i = 0) + 1
    row = lastRow + studentI + 1
    ws[f"A{row}"] = date
    ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
    ws[f"B{row}"] = studentName
    ws[f"C{row}"] = completeAvg
    ws[f"D{row}"] = performanceAvg
    # =KÖPRÜ(".. /images/student-based-reports/Ekran Görüntüsü-Png”> "Görüntü")
    ws[f"E{row}"] = f'=HYPERLINK(".{imgPath}", "Görüntü")'

    # sonraki satırdaki öğrenslye geçmek için
    # bir önceki sayfadaki tabloya geri dönüyoruz
    driver.back()

# 2 saniye bekletiyoruz
time.sleep(2)

# Sayfayı kapatıyoruz
driver.close()

# Excel dosyasını kaydet ve kapat
wb.save(xlPath)
wb.close()