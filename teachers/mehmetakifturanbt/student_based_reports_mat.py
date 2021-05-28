"""
ALGORİTMA
- Ekran görüntüleri için yoksa klasörü oluştur ("./images/student-based-reports")
- Excel dosyası eğer yoksa oluştur
- EBA'ya giriş yap
    ---------------------
    Karşılaşılan Sorunlar
    ---------------------
    - "EBA yükleniyor" mesajının gitmemesi (her sayfada karşılaşılabiliyor)
    - Sayfanın görünen kısmı yüklense bile, sekme başlığında yükleniyor resmi
        dönmeye devam ediyor ve sayfa yüklenmesi bitmiyor
- Raporlar > Çalışma Raporları > Öğrenci Bazlı sayfasına git
- Tablodaki öğrencileri gez. Her öğrenci için;
    - Son 10 çalışmayı kontrol et
    - Tamamlama ortalamasını hesap et
    - Performans ortalamasını hesap et
    - Excel' kaydet
"""
from selenium import webdriver
import settings
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def left_menu_is_loaded():
    try:
        wait.until(ec.visibility_of_element_located((By.ID, "vc-treeleftmenu")))
    except:
        print("Menünün yüklenmesi için çok bekledi. Sayfa yenileniyor...")
        driver.refresh()
        left_menu_is_loaded()

def login_e_devlet(tc, password):
    driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")
    if 'VCollabPlayer' in driver.current_url:
        return
    driver.find_element_by_xpath("//button[@title='edevlet girişi']").click()
    driver.find_element_by_css_selector("#tridField").send_keys(settings.tc)
    driver.find_element_by_xpath("//input[@id='egpField']").send_keys(settings.password)
    try:
        driver.find_element_by_css_selector("input[name='submitButton']").click()
        # ÖNEMLİ NOT: E-Devlet Girişinde Tek Kullanımlık Şifre Kullandığım İçin Kod Eklemesi Yaptım
        tek_kullanimlik_sifre = input("Tek Kullanımlık Şifrenizi Giriniz: ")
        driver.find_element_by_xpath("//input[@id='otpDogrulamaKodu']").send_keys(tek_kullanimlik_sifre)
        driver.find_element_by_css_selector("input[name='submitButton']").click()
    except:
        print("Sayfa 10 saniyede yüklenemedi. Sayfayı yeniliyorum!")
        login_e_devlet(tc, password)
    else:
        left_menu_is_loaded()


def login_mebbis(tc, password):
    driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")
    if 'VCollabPlayer' in driver.current_url:
        return
    driver.find_element_by_xpath("//button[@title='MEBBİS ile giriş']").click()
    driver.find_element_by_css_selector("#txtKullaniciAd").send_keys(settings.tc)
    driver.find_element_by_xpath("//input[@id='txtSifre']").send_keys(settings.password)
    try:
        guvenlik_kodu = input("Güvenlik Kodunu Giriniz: ")
        driver.find_element_by_xpath("//input[@id='txtGuvenlikKod']").send_keys(guvenlik_kodu)
        driver.find_element_by_css_selector("#btnGiris").click()
    except:
        print("Sayfa 10 saniyede yüklenemedi. Sayfayı yeniliyorum!")
        login_mebbis(tc, password)
    else:
        left_menu_is_loaded()


def table_is_loaded():
    try:
        return wait.until(ec.visibility_of_all_elements_located(
            (By.XPATH, "//div[@class='body-container']/div[@role='row']")
        ))
    except:
        print("Tablo yüklenemedi. Sayfayı yeniliyorum!")
        driver.refresh()
        return table_is_loaded()

def sinif_sube_ders_secimi(sinif, sube, ders):
    try:
        wait.until(ec.visibility_of_element_located(
            (By.XPATH, f"//option[contains(text(),'{sinif}. Sınıf')]")
        )).click()
        wait.until(ec.visibility_of_element_located(
            (By.XPATH, f"//option[contains(text(),'{sube} Şubesi')]")
        )).click()
        wait.until(ec.visibility_of_element_located(
            (By.XPATH, f"//option[@label='{ders}']")
        )).click()
        return f"{sinif}/{sube} Sınıfı {ders} Dersi İçin"
    except:
        print(f"{sinif}/{sube} Sınıfı ya da {ders} Dersi Bulunumadı.")

siniflar = ["10A","10B","10C","12A","12D"]

# Ekran görüntüleri için (eğer yoksa) klasör oluştur
# "./images/students-based-reports"
img_dir = "./images/students-based-reports"
if not os.path.exists(img_dir):
    os.mkdir(img_dir)

# Excel dosyası (eğer yoksa) oluştur
xl_path = "./excels/student-based-reports.xlsx"

if not os.path.exists(xl_path):
    wb = Workbook()
    ws = wb.active
    for i in range(len(siniflar)):
        ws = wb.create_sheet(title = f"{siniflar[i]}")
        ws.append(["Tarih", "Okul No", "Ad", "Soyad" , "Tamamlama", "Performans", "Ekran Görüntüsü"])

else:
    wb = load_workbook(xl_path)
    #ws = wb[2]


# tarayıcı nesnesi oluştur
driver = webdriver.Chrome(settings.driver_path)
driver.maximize_window()
# sayfanın yüklemesini çok beklememesi için 10 saniye beklemesini, yoksa hata vermesini belirliyoruz
driver.set_page_load_timeout(10)

wait = WebDriverWait(driver, timeout=10, poll_frequency=1)

# EBA'ya giriş
login_mebbis(settings.tc, settings.password)

# //div[@class='vc-lm-item-title '][normalize-space()='Raporlar']
# //div[@id='3e8510bd-2428-4fa8-12e1-42ea125d8ff2'] ID SÜREKLİ DEĞİŞTİĞİ İÇİN KULLANAMIYORUZ!! Önce bunu denemiştik.
# time.sleep(30)

# Raporlar Menüsüne Tıklama
reports_menu = wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
))
reports_menu.click()

# Çalışma Raporları Düğmesi
work_reports = wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[text()='Çalışma Raporları']")
))
work_reports.click()

table_is_loaded()

# Öğrenci Bazlı Sekmesi
student_based_link = wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")
))
student_based_link.click()

date = datetime.today()  # Günün tarihini al

# Sınıf 10 A , 10 B , 10 C, 12 A, 12 D seçebiliyorum.
for sinif in ("10","12"):
    for sube in ("A","B","C","D"):
        metin = sinif_sube_ders_secimi(sinif, sube, "Türk Dili ve Edebiyatı")
        students = table_is_loaded()
        student_count = len(students)
        if metin:
            print(f"{metin} {student_count} Adet Öğrenci Listelendi.")
        else:
            continue
        ws = wb[f"{sinif}{sube}"]
        last_row = ws.max_row  # Excel'de dolu olan son satırı al

        for student_i in range(student_count):
            students = table_is_loaded()
            students[student_i].click()

            works = table_is_loaded()
            student_name = driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text
            student_name = student_name.split(" ")
            okul_no = student_name.pop(0)
            student_surname = student_name.pop()
            student_name = " ".join(student_name)
            print(f"{okul_no} nolu {student_name} {student_surname} isimli öğrenciye ait {len(works)} adet çalışma listelendi.")

            completes = []
            performances = []

            for work_i in range(len(works)):
                complete = works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
                completes.append(int(complete.replace("%", "")))

                performance = works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
                if performance != "-":
                    performances.append(int(performance))

            complete_avg = sum(completes) / len(completes)

            performance_avg = "Performans Yok"
            if len(performances) > 0:
                performance_avg = sum(performances) / len(performances)


            print("*"*20)
            print("Öğrenci: ", student_name, student_surname)
            print("--- Tamamlama\t:", complete_avg)
            print("--- Performans\t:", performance_avg)
            print("*"*20)


            img_path = f"{img_dir}/{date.strftime('%Y%m%d')}_{okul_no}_{student_name}_{student_surname}.png"
            driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)


            row = last_row + student_i + 1
            ws[f"A{row}"] = date
            ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
            ws[f"B{row}"] = okul_no
            ws[f"C{row}"] = student_name
            ws[f"D{row}"] = student_surname
            ws[f"E{row}"] = complete_avg
            ws[f"F{row}"] = performance_avg
            ws[f"G{row}"] = f'=HYPERLINK(".{img_path}", "SS {student_name}")'



            driver.back()
            #break
        print("-"*50,"Şube işlemleri bitti")
    print("~"*60)

'''
# Benim uyarı verdiği için uyarı penceresindeki TAMAM düğmesine tıklama
driver.find_element_by_xpath("//a[normalize-space()='TAMAM']").click()

# Sayfam Menüsüne Tıklama
driver.find_element_by_xpath("//div[contains(@class,'vc-lm-item-title')][normalize-space()='Sayfam']").click()
time.sleep(5)

# Mesaj bölümüne settings dosyasında bulunan mesaj değişkenindeki metni yazdırma
driver.find_element_by_css_selector("textarea[placeholder='Ne paylaşmak istersin?']").send_keys(settings.mesaj)

# Mesajı hangi grupta paylaşacağımızı seçiyoruz. (Dropdown listeden seçim)
driver.find_element_by_xpath("//option[@label='Bilişim Zümresi']").click()
time.sleep(5)

# Paylaş Düğmesine Tıklama
driver.find_element_by_id("vc-PostButton").click()
time.sleep(5)

# Açılan Pop-Up Üzerindeki Paylaş Düğmesine Tıklama
driver.find_element_by_xpath("//a[contains(text(),'PAYLAŞ')]").click()
time.sleep(5)

# Paylaşılan Mesajı Silme
driver.find_element_by_xpath("//i[@class='fa fa-bars fa-lg']").click()
driver.find_element_by_xpath("//a[normalize-space()='Sil']").click()
time.sleep(1)
driver.find_element_by_xpath("//a[contains(text(),'SİL')]").click()
'''

# tarayıcı kapat
time.sleep(3)
driver.close()

wb.save(xl_path)
wb.close()
