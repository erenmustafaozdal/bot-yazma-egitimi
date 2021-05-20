"""
Linkedin sitesindeki bağlantıları kaydetmek
"""
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import settings
import os
from openpyxl import Workbook, load_workbook
import time


def page_scroll(second):
    SCROLL_PAUSE_TIME = second

    # Get scroll height
    last_height = driver.execute_script("return document.body.scrollHeight")

    while True:
        # Scroll down to bottom
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)

        # Calculate new scroll height and compare with last scroll height
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

# Ekran görüntüleri için (eğer yoksa) klasör oluştur
img_dir = "./images/linkedin_contacts"
if not os.path.exists(img_dir):
    os.mkdir(img_dir)

# Excel dosyası (eğer yoksa) oluştur
xl_path = "./excels/linkedin_contacts.xlsx"
if not os.path.exists(xl_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Contacts"
    ws.append(["Üyenin Adı", "Üyenin Çalışma Durumu", "Bağlantı Zamanı", "Ekran Görüntüsü"])

else:
    wb = load_workbook(xl_path)
    ws = wb["LinkedIn Contacts"]

driver = webdriver.Chrome(settings.driver_path)
wait = WebDriverWait(driver, timeout=10, poll_frequency=1)

driver.get("https://www.linkedin.com/")

driver.maximize_window()
time.sleep(1)

email = driver.find_element_by_xpath("//*[@id='session_key']")
password = driver.find_element_by_xpath("//*[@id='session_password']")

email.send_keys(settings.linkedin_mail)
password.send_keys(settings.linkedin_parola)
password.send_keys(Keys.ENTER)
time.sleep(2)

try:
    not_now_button = driver.find_element_by_xpath('//*[@id="remember-me-prompt__form-secondary"]/button')
    not_now_button.click()
except:
    pass

# Arama kutusu kullanma alıştırması
# search_bar = driver.find_element_by_xpath("//input[@placeholder='Arama Yap']")
# search_bar.send_keys("#python")
# search_bar.send_keys(Keys.RETURN)

driver.get("https://www.linkedin.com/mynetwork/invite-connect/connections/")

wait.until(ec.visibility_of_element_located(
            (By.XPATH, "//li-icon[@type='caret-filled-down-icon']//*[local-name()='svg']")
        )).click()

wait.until(ec.visibility_of_element_located(
            (By.XPATH, '//span[text()="Ad"]')
        )).click()

wait.until(ec.visibility_of_element_located(
            (By.XPATH, '//span[text()="Yeni Eklenen"]')
        )).click()

# for i in range(1, 20):
#     driver.execute_script('window.scrollTo(0,document.body.scrollHeight)')
#     time.sleep(3)

page_scroll(3)

contacts = driver.find_elements_by_class_name("mn-connection-card__details")
contactList=[]



last_row = ws.max_row  # Excel'de dolu olan son satırı al
sayac = 1
for contact in contacts:
    contactList.append(contact.text)
    bilgi = contact.text
    bilgi = bilgi.split("\n")
    ad_soyad = bilgi[1]
    meslek = bilgi[3]
    sure = bilgi[4][17:]
    img_path = f"{img_dir}/{ad_soyad}.png"
    driver.find_element_by_xpath(f"//li[@class='mn-connection-card artdeco-list ember-view'][{sayac}]").screenshot(img_path)
    last_row += 1
    sayac += 1
    ws[f"A{last_row}"] = ad_soyad
    ws[f"B{last_row}"] = meslek
    ws[f"C{last_row}"] = sure
    ws[f"D{last_row}"] = f'=HYPERLINK(".{img_path}", "SS_{ad_soyad}")'
    print(contact.text)

print(len(contactList))
with open("./txts/linkedin_contacts_mat.txt", "w", encoding="UTF-8") as file:
    for contact in contactList:
        file.write(contact + "\n")


time.sleep(5)
driver.quit()

wb.save(xl_path)
wb.close()














