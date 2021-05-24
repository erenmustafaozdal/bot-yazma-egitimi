from  selenium import webdriver
import settings
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec # as ile takma isim oluşturuldu
from selenium.webdriver.common.by import By
import os
# Öncelikle openpyxl paketi yüklenir: pip install openpyxl
from openpyxl import Workbook,load_workbook
from datetime import datetime

def left_menu_is_loaded(): #fonksiyon tanımladık
    try:
        #menü görünene kadar bekle
        wait.until(ec.visibility_of_element_located((By.ID, "vc-treeleftmenu")))
    except:
        print("Menünün yüklenmesi için çok bekledi. Sayfa yenileniyor....")
        driver.refresh()
        left_menu_is_loaded() # öz yinelemeli bir fonk oldu artık
        # biz bunu sonsuz döngüye sokuyoruz.
        # İstersek bir döngünün içine koyarak 3 defa çağırda diyebiliriz

 #tablo satırlarının görünmesini  bekleyen fonksiyon
def table_is_loaded():
    try:
     return wait.until(ec.visibility_of_all_elements_located(
     (By.XPATH, "//div[@class='col-container']")
     ))
    except:
        print("Tablo yüklenemedi. Sayfa yenileniyor.....")
        driver.refresh()
        return table_is_loaded()
# Ekran görüntüleri için yoksa klasörü oluştur
#  "./images/student-based-reports"
img_dir = "./images/student-based-reports"  #nokta projemizin ana dizini içndeki image
if not os.path.exists(img_dir):
    os.mkdir(img_dir) #yoksa oluştur make dir demek

def login(tc,password):   # giriş sayfası için fpnk tanımladık
    # Öğretmen giriş sayfasına git
   driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")

    # Eğer daha önceden giriş yapıldı ise fonksiyondan çık
    # 'VCollabPlayer' ifadesi EBA'ya giriş yaptıktan sonra
    # her URL'de olan ortak bir değerdir.
    # Geçerli URL içinde varlığı sorgulanarak,
    # giriş yapılıp yapılmadığı tespit edilir

   if ("VCollabPlayer" in driver.current_url):
        return # fonk burada durur aşağıya inmez

   # E-Devlet girişi tuşuna bas ve E-Devlet girişi sayfasına git
   driver.find_element_by_xpath("//button[@title='edevlet girişi']").click()

        # TC ve şifre yaz
   driver.find_element_by_xpath("//input[@id='tridField']").send_keys(settings.tc)
   driver.find_element_by_xpath("//input[@id='egpField']").send_keys(settings.password)

   # E-Devlet giriş formunu gönder.
   # Eğer sayfa yüklenemez veya başka bir hata alınırsa
   # giriş işlemini tekrar et.
   try:
    driver.find_element_by_xpath("//input[@name='submitButton']").click()
   except:
    print("Sayfa 20 saniyede yüklenemedi. Sayfa yenileniyor...")
    login(tc, password)
   else:
        # Eğer başarılı bir şekilde EBA'ya giriş
        # yapıldı ise sol menünün yüklenmesini bekle
        left_menu_is_loaded()
# Excel dosyası eğer yoksa oluştur
xl_path="./excels/student-based_report.xlsx"

if not os.path.exists(xl_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Öğrenci Bazlı Çalışma Raporları"
    ws.append(["Tarih", "Öğrenci", "Tamamlama", "Performans", "Ekran Görüntüsü"])
else:
    wb=load_workbook(xl_path)
    ws=wb["Öğrenci Bazlı Çalışma Raporları"]



# tarayıcı nesnesi oluştur
driver = webdriver.Chrome(settings.driver_path)
driver.maximize_window()

# sayfanın yüklemesini çok beklememesi için
# 20 saniye beklemesini, yoksa hata vermesini belirliyoruz
# hata verdiğinde bu hatayı yakalayıp (login fonksiyonunda)
# sayfanın yenilenmesini veya giriş işlemini tekrar etmesini sağlayacağız
driver.set_page_load_timeout(20)

# Çeşitli sayfa elemanlarını beklemek için
# WebDriverWait bekleme nesnesi oluşturulur
wait = WebDriverWait(driver, timeout=10, poll_frequency=1)

# EBA'ya giriş yap
login(settings.tc,settings.password)


# sol menüden raporlar bölümüne tıkla
reports_menu = wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']"))
).click()
work_reports = wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[text()='Çalışma Raporları']")
))
work_reports.click()

table_is_loaded()

#Öğrenci bazlıya tıklıyoruz
students_based_link=wait.until(ec.element_to_be_clickable(
    (By.XPATH,"//div[text()='ÖĞRENCİ BAZLI']"

)))

students_based_link.click()
#öğrencilerin olduğu tablo satırlarınıı
students= table_is_loaded()


students_count=len(students)
#günün tarihini al
date=datetime.today()
#excelde yazılmış en son satırı al
last_row=ws.max_row

for students_i in range(students_count):
    students=table_is_loaded()
    students[students_i].click()
    #çalışma satırlarını al
    works=table_is_loaded()
    student_name=driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text
    complates=[]
    performances=[]
    #çalışma satırlarını gez ve hesaplamalar yap
    for work_i in range(len(works)):
        complate= works[work_i].find_element_by_xpath('.//div[@id ="vcProgressBar"]//span').text
        complates.append(int(complate.replace("%","")))
        #performans al
        performance=works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
        if performance!="-":
            performances.append(int(performance))
    # tamamlama ortalamasını alalım
    complete_avg = sum(complates)/len(complates)
    # performans ortalamasını alalım
    performance_avg = "performans yok"  # varsayılan bir değer belirliyoruz.
    if len(performances) > 0:  # eğer performans değeri varsa ortlama hesap edilir
            performance_avg = sum(performances) / len(performances)
    # ekran görüntüsü alalım
    # Ekran Görüntüsü Adı Formatı: 20210508-Öğrenci Adı.png
    img_path = f"{img_dir}/{date.strftime('%Y%m%d')}-{student_name}.png"

    driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)

    print("*"*50)
    print("öğrenci:  ",student_name)
    print("Tamamlama: ",complete_avg)
    print("performans", performance_avg)
    #excel e yazdır
    #last_row =1 +(student_i=0)+1
    row=last_row+students_i+1
    ws[f"A{row}"] = date
    ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
    ws[f"B{row}"] = student_name
    ws[f"C{row}"] = complete_avg
    ws[f"D{row}"] = performance_avg
    # =KÖPRÜ("../images/student-based-reports/Ekran Görüntüsü.png";"Görüntü")
    ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'

    break

    driver.back()

time.sleep(2)
#tarayıcıyı kapat
driver.close()
#excel dosyasını kaydet ve kapat
wb.save(xl_path)
wb.close()
#//div[text()="Çalışma Raporları"] çalışma raporları için
#//div[@class='col-container'] tablo göründüğündeyı kontrol için
#//div[text()="ÖĞRENCİ BAZLI"] ÖĞRENCİ BAZLI BUTONA TIKLAMAK İÇİN KULLANACAĞIZ