
from selenium import webdriver
import settings
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import os
from openpyxl import Workbook, load_workbook


# SORUN ÇÖZÜMÜ: MEBBİS'e giriş yaptıkta sonra sayfa elemanlarının geç gelmesi.
# ÇÖZÜM: Duyurular paneli gelene kadar bekle
def mebbis_is_loaded():
    try:
        wait.until(ec.visibility_of_element_located((By.XPATH, "//div[contains(@class,'carousel')]")))
    except:
        print("Duyurular paneli için çok bekledi. Sayfa yenileniyor...")
        driver.refresh()
        mebbis_is_loaded()
    else:
        print("Mebbis duyuru paneli yüklendi.")

# e-Okul giriş sayfası açıldıktan sonra sol menünün yüklenmesini bekleme.
def eokul_navbar_is_loaded():
    try:
        wait.until(ec.visibility_of_all_elements_located((By.XPATH, "//div[contains(@class,'navbar')]")))
    except:
        print("Menünün yüklenmesi için çok bekledi. Sayfa yenileniyor...")
        driver.refresh()
        eokul_navbar_is_loaded()
    else:
        print("E-Okul navbar yüklendi.")


# Mebbis'e giriş yapma işlemini gerçekleştiren fonksiyon
def login(tc, password):
    # Mebbis giriş sayfasına git
    driver.get("https://mebbis.meb.gov.tr/default.aspx")
    driver.find_element_by_xpath("//input[@id='txtKullaniciAd']").send_keys(settings.tc)
    driver.find_element_by_xpath("//input[@id='txtSifre']").send_keys(settings.password)

    try:
        driver.find_element_by_xpath("//input[@id='btnGiris']").click()
    except:
        print("Sayfa 10 saniyede yüklenemedi. Sayfa yenileniyor...")
        login(tc, password)
    else:
        print("Mebbis Login başarılı.")
        # Eğer başarılı bir şekilde Mebbis'e giriş
        # yapıldı ise duyurular panelinin yüklenmesini bekle
        mebbis_is_loaded()


# Tablo satırlarının görünmesini bekleyen fonksiyon
def table_is_loaded():
    try:
        return wait.until(ec.visibility_of_all_elements_located(
            (By.XPATH, "//table[@id='dgListem']/tbody/tr")
        ))
    except:
        print("Liste yüklenemedi. Sayfa yenileniyor...")
        driver.get("https://e-okul.meb.gov.tr/OrtaOgretim/OKL/OOK07003.aspx")
        return "Yüklenemedi"
    else:
        print("Tablo yüklendi.")


# tarayıcı nesnesi oluştur
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.headless = True
options.add_argument(f"user-data-dir={os.getenv('USERPROFILE')}\\AppData\\Local\\Google\\Chrome\\User Data\\Default")
options.add_argument("--window-size=1920,1080")
driver = webdriver.Chrome(settings.driver_path,options=options)
driver.maximize_window()
# sayfanın yüklemesini çok beklememesi için
# 15 saniye beklemesini, yoksa hata vermesini belirliyoruz
# hata verdiğinde bu hatayı yakalayıp (login fonksiyonunda)
# sayfanın yenilenmesini veya giriş işlemini tekrar etmesini sağlayacağız
driver.set_page_load_timeout(15)

# Çeşitli sayfa elemanlarını beklemek için
# WebDriverWait bekleme nesnesi oluşturulur
wait = WebDriverWait(driver, timeout=15, poll_frequency=1)

# Mebbis'e giriş yapılır.
login(settings.tc, settings.password)
driver.find_element_by_xpath("//img[@src='images/uygulamaikon/e-okul.png']").click()
driver.find_elements_by_xpath("//a[contains(@title,'E-OKUL')]")[1].click()
# Yeni pencerede açılan E-Okul'a geçiş yapılıyor.
driver.switch_to.window(driver.window_handles[-1])
#e-okul sol menü gelene kadar bekleniyor.
eokul_navbar_is_loaded()
# Ders notu girişi ekranına kadar gerekli menü tıklamaları yapılıyor.
driver.find_element_by_xpath("//span[contains(text(),'Ortaöğretim Kurum İşlemleri')]").click()
time.sleep(0.5)
driver.find_element_by_xpath("//span[contains(text(),'Not İşlemleri')]").click()
time.sleep(0.5)
driver.find_element_by_xpath("//a[contains(text(),'Ders Notu Girişi')]").click()

while True:
    #baslik değişkeni excel sheetine vereceğimiz title olacak
    baslik=""
    #Öğretmene tanımlı şubeler alınıyor.
    subeler = driver.find_elements_by_xpath("//select[@id='Us_SinifSube1_ddlSinifSube']/option")
    #Öğretmene hangi sınıfın listesini istediği soruluyor ve o şube tıklanıyor.
    i = 1
    print("Liste almak istediğiniz şubeyi seçiniz:")
    for sube in subeler:
        print(str(i), "- ", sube.text)
        i = i + 1
    driver.minimize_window()
    secim = int(input())
    baslik=baslik+subeler[secim - 1].text
    driver.maximize_window()
    subeler[secim - 1].click()
    time.sleep(0.5)
    #Öğretmene tanımlı dersler alınıyor.
    dersler = driver.find_elements_by_xpath("//select[@id='ddlDersler']/option")
    #Öğretmene hangi dersin listesini istediği soruluyor ve o ders tıklanıyor.
    i = 0
    print("Liste almak istediğiniz dersi seçiniz:")
    for ders in dersler:
        if len(ders.text)>1:
            print(str(i), "- ", ders.text)
        i = i + 1
    driver.minimize_window()
    secim = int(input())
    baslik=baslik+dersler[secim].text
    driver.maximize_window()
    dersler[secim].click()
    time.sleep(0.5)
    #Seçilen şube ve dersin istesini almak için Listele düğmesi tıklanıyor.
    driver.find_element_by_id("btnListele").click()
    table_is_loaded()
    #liste yüklenemezse while döngüsünün başına dönüyor.
    if table_is_loaded()=="Yüklenemedi":
        print("Sayfa yenileniyor")
    else:
        print("Liste yüklendi.")
        break

# Öğrencilerin olduğu tablo satırlarını al
students = table_is_loaded()
student_count = len(students)
#başlık içindeki gereksiz karakterler temizleniyor ve kısaltılıyor.
geresizkarakter="-./() "
for kar in geresizkarakter: baslik=baslik.replace(kar,"")
baslik=baslik.replace("Sınıf","")
baslik=baslik.replace("Şubesi","")
if "SEÇMELİ" in baslik:
    baslik=baslik.replace("SEÇMELİ","SEÇ")
# Excel dosyası eğer yoksa oluştur
xl_path = "./excels/Not_Listeleri.xlsx"
if not os.path.exists(xl_path):
    wb = Workbook()
    ws = wb.active
    ws.title = baslik
else:
    wb = load_workbook(xl_path)
    #Eğer daha önce bu sınıf ve dersin listesi oluşturulmuşsa eskisi siliniyor.
    if baslik in wb.sheetnames:
        print("Sınıf Listesi daha önce oluşturulmuş.Eskisi siliniyor.")
        ws = wb[baslik]
        wb.remove(ws)
        ws=wb.create_sheet(baslik)
    else:
        print("Sınıf Listesi oluşturuluyor.")
        ws=wb.create_sheet(baslik)
#Öğrenci ve not bilgileri excel dosyasına yazılıyor.
for student_i in range(student_count):
    print(students[student_i].find_element_by_xpath("td[1]").text,"nolu öğrenci alındı.")
    row =student_i + 1
    ws[f"A{row}"] = students[student_i].find_element_by_xpath("td[1]").text
    ws[f"B{row}"] = students[student_i].find_element_by_xpath("td[2]").text
    ws[f"C{row}"] = students[student_i].find_element_by_xpath("td[3]").text
    ws[f"D{row}"] = students[student_i].find_element_by_xpath("td[4]").text
    ws[f"E{row}"] = students[student_i].find_element_by_xpath("td[5]").text
    ws[f"F{row}"] = students[student_i].find_element_by_xpath("td[6]").text
    ws[f"G{row}"] = students[student_i].find_element_by_xpath("td[7]").text
    ws[f"H{row}"] = students[student_i].find_element_by_xpath("td[8]").text
    ws[f"I{row}"] = students[student_i].find_element_by_xpath("td[9]").text
    ws[f"J{row}"] = students[student_i].find_element_by_xpath("td[10]").text
    ws[f"K{row}"] = students[student_i].find_element_by_xpath("td[11]").text
    ws[f"L{row}"] = students[student_i].find_element_by_xpath("td[12]").text
    ws[f"M{row}"] = students[student_i].find_element_by_xpath("td[13]").text

# tarayıcı kapat
driver.close()
driver.quit()
# # Excel dosyasını kaydet ve kapat
wb.save(xl_path)
wb.close()
