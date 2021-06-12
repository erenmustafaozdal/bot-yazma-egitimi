from selenium import webdriver
import settings
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def altincisinif():
    def left_menu_is_loaded():
        try:
            wait.until(ec.visibility_of_element_located((By.ID, "vc-treeleftmenu")))
        except:
            print("Menünün yüklenmesi için çok bekledi. Sayfa yenileniyor...")
            driver.refresh()
            left_menu_is_loaded()
    def login(tc, password):
        driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")
        if 'VCollabPlayer' in driver.current_url:
            return
        driver.find_element_by_xpath("//button[@title='edevlet girişi']").click()
        driver.find_element_by_css_selector("#tridField").send_keys(settings.tc)
        driver.find_element_by_id("egpField").send_keys(settings.password)
        try:
            driver.find_element_by_xpath("//input[@name='submitButton']").click()
        except:
            print("Sayfa 20 saniyede yüklenemedi. Sayfa yenileniyor...")
            login(tc, password)
        else:
            left_menu_is_loaded()
    def table_is_loaded():
        try:
            return wait.until(ec.visibility_of_all_elements_located(
                (By.XPATH, "//div[@class='body-container']/div[@role='row']")
            ))
        except:
            print("Tablo yüklenemedi. Sayfa yenileniyor...")
            driver.refresh()
            return table_is_loaded()
    img_dir = "./images/student-based-reports"
    if not os.path.exists(img_dir):
        os.mkdir(img_dir)
    xl_path = "./excels/student-based_reports.xlsx"
    if not os.path.exists(xl_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Öğrenci Bazlı Çalışma Raporları"
        ws.append(["Tarih", "Öğrenci", "Tamamlama", "Performans", "Ekran Görüntüsü"])
    else:
        wb = load_workbook(xl_path)
        ws = wb["Öğrenci Bazlı Çalışma Raporları"]
    driver = webdriver.Chrome(settings.driver_path)
    driver.maximize_window()
    driver.set_page_load_timeout(20)
    wait = WebDriverWait(driver, timeout=10, poll_frequency=1)
    login(settings.tc, settings.password)
    reports_menu = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
    ))
    reports_menu.click()
    work_reports = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='Çalışma Raporları']")
    ))
    work_reports.click()
    table_is_loaded()
    student_based_link = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")
    ))
    student_based_link.click()
    students = table_is_loaded()
    student_count = len(students)
    date = datetime.today()
    last_row = ws.max_row
    for student_i in range(student_count):
        students = table_is_loaded()
        students[student_i].click()
        works = table_is_loaded()
        student_name = driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text
        completes = []
        performances = []
        for work_i in range(10):
            complete = works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
            completes.append(int(complete.replace("%", "")))
            performance = works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
            if performance != '-':
                performances.append(int(performance))
        complete_avg = sum(completes) / len(completes)
        performance_avg = "performans yok"
        if len(performances) > 0:
            performance_avg = sum(performances) / len(performances)
        img_path = f"{img_dir}/{date.strftime('%Y%m%d')}-{student_name}.png"
        driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)
        print("*" * 50)
        print("Öğrenci: ", student_name)
        print("--- Tamamlama:", complete_avg)
        print("--- Performans:", performance_avg)
        row = last_row + student_i + 1
        ws[f"A{row}"] = date
        ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
        ws[f"B{row}"] = student_name
        ws[f"C{row}"] = complete_avg
        ws[f"D{row}"] = performance_avg
        ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'
        driver.back()
    time.sleep(2)
    driver.close()
    wb.save(xl_path)
    wb.close()

def yedincisinif():
    def left_menu_is_loaded():
        try:
            wait.until(ec.visibility_of_element_located((By.ID, "vc-treeleftmenu")))
        except:
            print("Menünün yüklenmesi için çok bekledi. Sayfa yenileniyor...")
            driver.refresh()
            left_menu_is_loaded()
    def login(tc, password):
        driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")
        if 'VCollabPlayer' in driver.current_url:
            return
        driver.find_element_by_xpath("//button[@title='edevlet girişi']").click()
        driver.find_element_by_css_selector("#tridField").send_keys(settings.tc)
        driver.find_element_by_id("egpField").send_keys(settings.password)
        try:
            driver.find_element_by_xpath("//input[@name='submitButton']").click()
        except:
            print("Sayfa 20 saniyede yüklenemedi. Sayfa yenileniyor...")
            login(tc, password)
        else:
            left_menu_is_loaded()
    def table_is_loaded():
        try:
            return wait.until(ec.visibility_of_all_elements_located(
                (By.XPATH, "//div[@class='body-container']/div[@role='row']")
            ))
        except:
            print("Tablo yüklenemedi. Sayfa yenileniyor...")
            driver.refresh()
            return table_is_loaded()
    img_dir = "./images/student-based-reports"
    if not os.path.exists(img_dir):
        os.mkdir(img_dir)
    xl_path = "./excels/student-based_reports.xlsx"
    if not os.path.exists(xl_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Öğrenci Bazlı Çalışma Raporları"
        ws.append(["Tarih", "Öğrenci", "Tamamlama", "Performans", "Ekran Görüntüsü"])
    else:
        wb = load_workbook(xl_path)
        ws = wb["Öğrenci Bazlı Çalışma Raporları"]
    driver = webdriver.Chrome(settings.driver_path)
    driver.maximize_window()
    driver.set_page_load_timeout(20)
    wait = WebDriverWait(driver, timeout=10, poll_frequency=1)
    login(settings.tc, settings.password)
    reports_menu = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
    ))
    reports_menu.click()
    work_reports = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='Çalışma Raporları']")
    ))
    work_reports.click()
    table_is_loaded()
    student_based_link = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")
    ))
    student_based_link.click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='7. Sınıf']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='Fen Bilimleri']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='Tüm Dersler']").click()
    time.sleep(5)
    students = table_is_loaded()
    student_count = len(students)
    date = datetime.today()
    last_row = ws.max_row
    for student_i in range(student_count):
        students = table_is_loaded()
        students[student_i].click()
        works = table_is_loaded()
        student_name = driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text
        completes = []
        performances = []
        for work_i in range(10):
            complete = works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
            completes.append(int(complete.replace("%", "")))
            performance = works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
            if performance != '-':
                performances.append(int(performance))
        complete_avg = sum(completes) / len(completes)
        performance_avg = "performans yok"
        if len(performances) > 0:
            performance_avg = sum(performances) / len(performances)
        img_path = f"{img_dir}/{date.strftime('%Y%m%d')}-{student_name}.png"
        driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)
        print("*" * 50)
        print("Öğrenci: ", student_name)
        print("--- Tamamlama:", complete_avg)
        print("--- Performans:", performance_avg)
        row = last_row + student_i + 1
        ws[f"A{row}"] = date
        ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
        ws[f"B{row}"] = student_name
        ws[f"C{row}"] = complete_avg
        ws[f"D{row}"] = performance_avg
        ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'
        driver.back()
    time.sleep(2)
    driver.close()
    wb.save(xl_path)
    wb.close()

def sekizincisinif():
    def left_menu_is_loaded():
        try:
            wait.until(ec.visibility_of_element_located((By.ID, "vc-treeleftmenu")))
        except:
            print("Menünün yüklenmesi için çok bekledi. Sayfa yenileniyor...")
            driver.refresh()
            left_menu_is_loaded()
    def login(tc, password):
        driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")
        if 'VCollabPlayer' in driver.current_url:
            return
        driver.find_element_by_xpath("//button[@title='edevlet girişi']").click()
        driver.find_element_by_css_selector("#tridField").send_keys(settings.tc)
        driver.find_element_by_id("egpField").send_keys(settings.password)
        try:
            driver.find_element_by_xpath("//input[@name='submitButton']").click()
        except:
            print("Sayfa 20 saniyede yüklenemedi. Sayfa yenileniyor...")
            login(tc, password)
        else:
            left_menu_is_loaded()
    def table_is_loaded():
        try:
            return wait.until(ec.visibility_of_all_elements_located(
                (By.XPATH, "//div[@class='body-container']/div[@role='row']")
            ))
        except:
            print("Tablo yüklenemedi. Sayfa yenileniyor...")
            driver.refresh()
            return table_is_loaded()
    img_dir = "./images/student-based-reports"
    if not os.path.exists(img_dir):
        os.mkdir(img_dir)
    xl_path = "./excels/student-based_reports.xlsx"
    if not os.path.exists(xl_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Öğrenci Bazlı Çalışma Raporları"
        ws.append(["Tarih", "Öğrenci", "Tamamlama", "Performans", "Ekran Görüntüsü"])
    else:
        wb = load_workbook(xl_path)
        ws = wb["Öğrenci Bazlı Çalışma Raporları"]
    driver = webdriver.Chrome(settings.driver_path)
    driver.maximize_window()
    driver.set_page_load_timeout(20)
    wait = WebDriverWait(driver, timeout=10, poll_frequency=1)
    login(settings.tc, settings.password)
    reports_menu = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
    ))
    reports_menu.click()
    work_reports = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='Çalışma Raporları']")
    ))
    work_reports.click()
    table_is_loaded()
    student_based_link = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")
    ))
    student_based_link.click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='8. Sınıf']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='Fen Bilimleri']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='Tüm Dersler']").click()
    students = table_is_loaded()
    student_count = len(students)
    date = datetime.today()
    last_row = ws.max_row
    for student_i in range(student_count):
        students = table_is_loaded()
        students[student_i].click()
        works = table_is_loaded()
        student_name = driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text
        completes = []
        performances = []
        for work_i in range(10):
            complete = works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
            completes.append(int(complete.replace("%", "")))
            performance = works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
            if performance != '-':
                performances.append(int(performance))
        complete_avg = sum(completes) / len(completes)
        performance_avg = "performans yok"
        if len(performances) > 0:
            performance_avg = sum(performances) / len(performances)
        img_path = f"{img_dir}/{date.strftime('%Y%m%d')}-{student_name}.png"
        driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)
        print("*" * 50)
        print("Öğrenci: ", student_name)
        print("--- Tamamlama:", complete_avg)
        print("--- Performans:", performance_avg)
        row = last_row + student_i + 1
        ws[f"A{row}"] = date
        ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
        ws[f"B{row}"] = student_name
        ws[f"C{row}"] = complete_avg
        ws[f"D{row}"] = performance_avg
        ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'
        driver.back()
    time.sleep(2)
    driver.close()
    wb.save(xl_path)
    wb.close()

def tumsiniflar():
    def left_menu_is_loaded():
        try:
            wait.until(ec.visibility_of_element_located((By.ID, "vc-treeleftmenu")))
        except:
            print("Menünün yüklenmesi için çok bekledi. Sayfa yenileniyor...")
            driver.refresh()
            left_menu_is_loaded()
    def login(tc, password):
        driver.get("https://giris.eba.gov.tr/EBA_GIRIS/teacher.jsp")
        if 'VCollabPlayer' in driver.current_url:
            return
        driver.find_element_by_xpath("//button[@title='edevlet girişi']").click()
        driver.find_element_by_css_selector("#tridField").send_keys(settings.tc)
        driver.find_element_by_id("egpField").send_keys(settings.password)
        try:
            driver.find_element_by_xpath("//input[@name='submitButton']").click()
        except:
            print("Sayfa 20 saniyede yüklenemedi. Sayfa yenileniyor...")
            login(tc, password)
        else:
            left_menu_is_loaded()
    def table_is_loaded():
        try:
            return wait.until(ec.visibility_of_all_elements_located(
                (By.XPATH, "//div[@class='body-container']/div[@role='row']")
            ))
        except:
            print("Tablo yüklenemedi. Sayfa yenileniyor...")
            driver.refresh()
            return table_is_loaded()
    img_dir = "./images/student-based-reports"
    if not os.path.exists(img_dir):
        os.mkdir(img_dir)
    xl_path = "./excels/student-based_reports.xlsx"
    if not os.path.exists(xl_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Öğrenci Bazlı Çalışma Raporları"
        ws.append(["Tarih", "Öğrenci", "Tamamlama", "Performans", "Ekran Görüntüsü"])
    else:
        wb = load_workbook(xl_path)
        ws = wb["Öğrenci Bazlı Çalışma Raporları"]
    driver = webdriver.Chrome(settings.driver_path)
    driver.maximize_window()
    driver.set_page_load_timeout(20)
    wait = WebDriverWait(driver, timeout=10, poll_frequency=1)
    login(settings.tc, settings.password)
    reports_menu = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
    ))
    reports_menu.click()
    work_reports = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='Çalışma Raporları']")
    ))
    work_reports.click()
    table_is_loaded()
    student_based_link = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")
    ))
    student_based_link.click()
    students = table_is_loaded()
    student_count = len(students)
    date = datetime.today()
    last_row = ws.max_row
    for student_i in range(student_count):
        students = table_is_loaded()
        students[student_i].click()
        works = table_is_loaded()
        student_name = driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text
        completes = []
        performances = []
        for work_i in range(10):
            complete = works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
            completes.append(int(complete.replace("%", "")))
            performance = works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
            if performance != '-':
                performances.append(int(performance))
        complete_avg = sum(completes) / len(completes)
        performance_avg = "performans yok"
        if len(performances) > 0:
            performance_avg = sum(performances) / len(performances)
        img_path = f"{img_dir}/{date.strftime('%Y%m%d')}-{student_name}.png"
        driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)
        print("*" * 50)
        print("Öğrenci: ", student_name)
        print("--- Tamamlama:", complete_avg)
        print("--- Performans:", performance_avg)
        row = last_row + student_i + 1
        ws[f"A{row}"] = date
        ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
        ws[f"B{row}"] = student_name
        ws[f"C{row}"] = complete_avg
        ws[f"D{row}"] = performance_avg
        ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'
        driver.back()
    time.sleep(2)
    wb.save(xl_path)
    wb.close()
    reports_menu = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
    ))
    reports_menu.click()
    work_reports = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='Çalışma Raporları']")
    ))
    work_reports.click()
    table_is_loaded()
    student_based_link = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")
    ))
    student_based_link.click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='7. Sınıf']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='Fen Bilimleri']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='Tüm Dersler']").click()
    time.sleep(5)
    students = table_is_loaded()
    student_count = len(students)
    date = datetime.today()
    last_row = ws.max_row
    for student_i in range(student_count):
        students = table_is_loaded()
        students[student_i].click()
        works = table_is_loaded()
        student_name = driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text
        completes = []
        performances = []
        for work_i in range(10):
            complete = works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
            completes.append(int(complete.replace("%", "")))
            performance = works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
            if performance != '-':
                performances.append(int(performance))
        complete_avg = sum(completes) / len(completes)
        performance_avg = "performans yok"
        if len(performances) > 0:
            performance_avg = sum(performances) / len(performances)
        img_path = f"{img_dir}/{date.strftime('%Y%m%d')}-{student_name}.png"
        driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)
        print("*" * 50)
        print("Öğrenci: ", student_name)
        print("--- Tamamlama:", complete_avg)
        print("--- Performans:", performance_avg)
        row = last_row + student_i + 1
        ws[f"A{row}"] = date
        ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
        ws[f"B{row}"] = student_name
        ws[f"C{row}"] = complete_avg
        ws[f"D{row}"] = performance_avg
        ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'
        driver.back()
    time.sleep(2)
    wb.save(xl_path)
    wb.close()
    reports_menu = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
    ))
    reports_menu.click()
    work_reports = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='Çalışma Raporları']")
    ))
    work_reports.click()
    table_is_loaded()
    student_based_link = wait.until(ec.element_to_be_clickable(
        (By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")
    ))
    student_based_link.click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='8. Sınıf']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='Fen Bilimleri']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//option[@label='Tüm Dersler']").click()
    students = table_is_loaded()
    student_count = len(students)
    date = datetime.today()
    last_row = ws.max_row
    for student_i in range(student_count):
        students = table_is_loaded()
        students[student_i].click()
        works = table_is_loaded()
        student_name = driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text
        completes = []
        performances = []
        for work_i in range(10):
            complete = works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
            completes.append(int(complete.replace("%", "")))
            performance = works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
            if performance != '-':
                performances.append(int(performance))
        complete_avg = sum(completes) / len(completes)
        performance_avg = "performans yok"
        if len(performances) > 0:
            performance_avg = sum(performances) / len(performances)
        img_path = f"{img_dir}/{date.strftime('%Y%m%d')}-{student_name}.png"
        driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)
        print("*" * 50)
        print("Öğrenci: ", student_name)
        print("--- Tamamlama:", complete_avg)
        print("--- Performans:", performance_avg)
        row = last_row + student_i + 1
        ws[f"A{row}"] = date
        ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
        ws[f"B{row}"] = student_name
        ws[f"C{row}"] = complete_avg
        ws[f"D{row}"] = performance_avg
        ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'
        driver.back()
    time.sleep(2)
    driver.close()
    wb.save(xl_path)
    wb.close()
while True:
    secim = int(input('''
    EBA Öğrenci Çalışma Raporlarına Bakma
    ---------------------------------------------------------------
    1 Altıncı Sınıf
    2 Yedinci Sınıf
    3 Sekizinci Sınıf 
    4 Tüm Sınıflar
    0 ÇIKIŞ
    
    Hangi Sınıfı Kontrol Etmek İstiyorsunuz:'''))

    if secim == 0:
        break
    if secim == 1:
        altincisinif()
    if secim == 2:
        yedincisinif()
    if secim == 3:
        sekizincisinif()
    if secim == 4:
        tumsiniflar()
