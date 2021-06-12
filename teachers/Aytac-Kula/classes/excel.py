from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import win32com.client as win32
import os


class Excel:
    """
    Excel çalışma kitabındaki istenen bir çalışma sayfanın (sekmenin) içindeki verileri başlık (headers) ve veriler (datas) şekline getirir. Bu verileri dictionary yapısıyla anahtar (key), değer (value) şeklinde erişilmesini sağlar.
    """

    def __init__(self, path, worksheet=None, data_only=False):
        """
        Yapılandırıcı metot

        :param path: excel dosyasının göreceli veya tam yolu
        :param worksheet: çalışma sayfasının adı veya indis (index) numarası. Eğer ayarlanmazsa atif olan çalışma sayfası seçilir
        """

        # data_only modda formüllerin değerledirilmemesi sorununu çözmek için
        # recalculation_formula metodu ile yeniden formül hesaplama işlemi
        # yapılıp yapılmadığını tutar. Eğer yeniden hesaplama yapıldıysa
        # tekrar yapılmasını engeller
        self.is_recalculate = False
        self.data_only = data_only

        # Excel dosyası yolu
        self.path = path

        # gerekiyorsa formülleri yeniden hesapla
        self.recalculation_formula()

        # çalışma kitabı
        self.wb = load_workbook(path, data_only=data_only)

        # çalışma sayfası
        if worksheet is None:
            self.ws = self.wb.active
        elif type(worksheet) is int:
            self.ws = self.wb.worksheets[worksheet]
        else:
            self.ws = self.wb[worksheet]

        self.headers = {}  # çalışma sayfası başlıkları
        self.datas = [] # çalışma sayfası verileri

        # başlıkları belirle
        self.set_headers()

    def set_headers(self):
        """
        Çalışma sayfasının başlıklarını aşağıdaki yapıya benzer şekilde belirler

        {
            "1. Sütun Başlığı" : {"row":1, "column":1},
            "2. Sütun Başlığı" : {"row":1, "column":2},
        }
        """
        header_row = self.ws[1]
        for header in header_row:
            self.headers[header.value] = {
                'row': header.row,
                'column': get_column_letter(header.column),
                'column_num': header.column
            }

    def get_datas(self):
        """
        Çalışma sayfası içindeki verileri aşağıdaki yapıdı düzenler ve geri döndürür.

        [
            {
                "1. Sütun Başlığı" : "1. satır 1. sütun verisi",
                "2. Sütun Başlığı" : "1. satır 2. sütun verisi",
                "row" : 1"
            },
            {
                "1. Sütun Başlığı" : "2. satır 1. sütun verisi",
                "2. Sütun Başlığı" : "2. satır 2. sütun verisi",
                "row" : 2"
            }
        ]

        :return: çalışma sayfasındaki başlık hariç satırlar
        """

        # eğer daha önce veriler alındı ise döndür
        if len(self.datas) > 0:
            return self.datas

        # eğer daha önce veriler alınmadı ise al ve düzenle
        for row in self.ws.rows:
            row_number = row[0].row
            # eğer başlık satırı ise geç
            if row_number == 1:
                continue

            row_data = {'row' : row_number}
            for header in self.headers:
                column = self.headers[header]['column_num']
                row_data[header] = row[column-1].value

            self.datas.append(row_data)

        print(f"'{self.ws.title}' içindeki veriler alındı.")

        return self.datas

    def update_range(self, worksheet_range, data, skip_first=False):
        """
        Belirli bir sütundaki değerleri baştan sona üzerine yazarak günceller

        :param worksheet_range: çalışma sayfası aralığı
        :param data: çalışma sayfası aralığını güncelleyecek veriler
        :param skip_first: ilk hücre atlansın mı? (sütun güncellenirken başlık geçilebilir)
        """

        data_count = len(data)
        for index, cell in enumerate(self.ws[worksheet_range]):
            # ilk hücreyse ve ilk hücre atlanacaksa; atla
            if index == 0 and skip_first:
                continue

            # eğer ilk hücre atlanıldıysa; data index'te kayma olacaktır
            # bunu aşmak için ilk hücre atlandıysa index'i 1 düşürüyoruz
            data_index = index if not skip_first else index - 1

            # eğer data bitti ise döngüden çık
            if data_index == data_count:
                break

            cell.value = data[data_index]

        print(f"'{self.ws.title}' içinde '{worksheet_range}' aralığı güncellendi.")

    def recalculation_formula(self):
        """
        Formüllerin yeniden hesaplanıp Excel ön belleğine alınmasını sağlar.
        Bir Excel dosyası 'data_only' modunda açıldığında; eğer Excel ön belleği yoksa,
        formülleri None olarak döndürüyor. Bunun önüne geçmek için dosyayı
        Excel uygulaması ile açıp, kaydedip kapatıyoruz.
        """

        # eğer yeniden formül hesaplama daha önceden yapıldıysa çık
        if self.is_recalculate or not self.data_only:
            return

        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workbook = excel.Workbooks.Open(os.path.abspath(self.path))
        workbook.Save()
        workbook.Close()
        excel.Quit()

        self.is_recalculate = True

    def save(self):
        """
        Excel çalışma kitabını kaydeder
        """
        self.wb.save(self.path)

        print(f"'{self.ws.title}' çalışma sayfasına sahip Excel dosyası kaydedildi.")

    def __del__(self):
        """
        Nesne silindiğinde Excel çalışma kitabını kapatır
        """
        self.wb.close()

        print(f"'{self.ws.title}' çalışma sayfasına sahip Excel dosyası kapatıldı.")
