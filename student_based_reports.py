"""
ALGORİTMA
- Ekran görüntüleri için yoksa klasörü oluştur ("./images/student-based-reports")
- Excel dosyası eğer yoksa oluştur
- EBA'ya giriş yap
    ---------------------
    Karşılaşılan Sorunlar
    ---------------------
    - "EBA yükleniyor" mesajının gitmemesi (her sayfada karşılaşılabiliyor)
    - Sayfanın görünen kısmı yüklense bile, sekme başlığında yükleniyor resmi
        dönmeye devam ediyor ve sayfa yüklenmesi bitmiyor
- Raporlar > Çalışma Raporları > Öğrenci Bazlı sayfasına git
- Tablodaki öğrencileri gez. Her öğrenci için;
    - Son 10 çalışmayı kontrol et
    - Tamamlama ortalamasını hesap et
    - Performans ortalamasını hesap et
    - Excel' kaydet
"""
import settings
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import os
# Öncelikle openpyxl paketi yüklenir: pip install openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
from classes.eba import EBA
from classes.browser import Browser


# Ekran görüntüleri için yoksa klasörü oluştur
#  "./images/student-based-reports"
img_dir = "./images/student-based-reports"
if not os.path.exists(img_dir):
    os.mkdir(img_dir)

# Excel dosyası eğer yoksa oluştur
xl_path = "./excels/student-based_reports.xlsx"
if not os.path.exists(xl_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Öğrenci Bazlı Çalışma Raporları"
    ws.append(["Tarih", "Öğrenci", "Tamamlama", "Performans", "Ekran Görüntüsü"])
else:
    wb = load_workbook(xl_path)
    ws = wb["Öğrenci Bazlı Çalışma Raporları"]


# tarayıcı nesnesi oluştur
# Çeşitli sayfa elemanlarını beklemek için
# WebDriverWait bekleme nesnesi oluşturulur
wait = WebDriverWait(driver, timeout=10, poll_frequency=1)
driver = Browser(settings.driver_path).get()

# EBA nesnesi oluşturulur
eba = EBA(driver)
# EBA'ya giriş yapılır
eba.login(settings.tc, settings.password)


# Sol menüden 'Raporlar' menüsüne tıkla
reports_menu = wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[@class='vc-lm-item-title '][normalize-space()='Raporlar']")
))
reports_menu.click()


# 'Raporlar' sayfasında 'Çalışma Raporları' bağlantısına tıkla.
work_reports = wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[text()='Çalışma Raporları']")
))
work_reports.click()

# Bu sayfanın yüklendiğinden emin olmak için tablo satırları yüklenene kadar bekle
eba.table_is_loaded()


# 'Çalışma Raporları' sayfasında 'ÖĞRENCİ BAZLI' bağlantısına tıkla.
student_based_link = wait.until(ec.element_to_be_clickable(
    (By.XPATH, "//div[text()='ÖĞRENCİ BAZLI']")
))
student_based_link.click()

# Bu sayfanın yüklendiğinden emin olmak için tablo satırları yüklenene kadar bekle.
# Öğrencilerin olduğu tablo satırlarını al
students = eba.table_is_loaded()


# Öğrencilerin satırlarını döngüye al.
# 1. Öğrencinin adını al.
# 2. Son 10 çalışmanın tamamlanma yüzdelerini bir diziye aktar.
# 3. Son 10 çalışmanın performanslarını bir diziye aktar.
# 4. Ortamala tamamlamayı hesap et.
# 5. Ortalama performansı hesap et.
# 6. Ekran görüntüsünü kaydet.
# 7. Verileri ekrana yazdır.
# TODO: 8. Verileri Excel dosyasına yazdır.
student_count = len(students)

# Günün tarihini al
date = datetime.today()
# Excel'de yazılmış en son satırı al
last_row = ws.max_row
for student_i in range(student_count):
    # Sayfa her yenilendiğinde elemanları baştan oluşturulur.
    # Yukarıda "students" değişkenine satırlar aktarılsa bile;
    # ilk öğrenci kontrol edilip, döngünün en altında
    # önceki sayfaya geri döndüğünde; sonraki öğrencinin satırı
    # yeniden oluşturulduğu için tıklama yapılamıyordu.
    # Bu sebeple döngü her döndüğünde satırlar tekrar alınır ve sıradaki satır tıklanır.
    students = eba.table_is_loaded()
    students[student_i].click()

    # çalışma satırlarını al
    works = eba.table_is_loaded()

    student_name = driver.find_element_by_xpath("//div[@class='vc-font-size-x-large m-l-sm ng-binding']").text

    # çalışma satırlarını gez ve hesaplamalar yap
    completes = []  # tamamlama yüzdeleri bu listeye atılarak hesaplanacak
    performances = []  # performanslar bu listeye atılarak hesaplanacak

    # eğer bütün çalışmalar gezilecekse "for work in works:" şeklinde
    # döngü oluşturulabilir. Bu örnekte son 10 çalışma kontrol ediliyor.
    for work_i in range(10):
        # tamamlanma alınır
        complete = works[work_i].find_element_by_xpath(".//div[@id='vcProgressBar']//span").text
        completes.append(int(complete.replace("%", "")))

        # performans alınır
        performance = works[work_i].find_element_by_xpath(".//div[@id='multiColouredProgress']//span").text
        if performance != '-':
            performances.append(int(performance))

    # tamamlama ortalamasını alalım
    complete_avg = sum(completes) / len(completes)
    # performans ortalamasını alalım
    performance_avg = "performans yok"  # varsayılan bir değer belirliyoruz.
    if len(performances) > 0:  # eğer performans değeri varsa ortlama hesap edilir
        performance_avg = sum(performances) / len(performances)

    # ekran görüntüsü alalım
    # Ekran Görüntüsü Adı Formatı: 20210508-Öğrenci Adı.png
    img_path = f"{img_dir}/{date.strftime('%Y%m%d')}-{student_name}.png"
    driver.find_element_by_xpath("//div[@class='vc-layout-view-content-padding']").screenshot(img_path)

    # ekrana yazdır
    print("*"*50)
    print("Öğrenci: ", student_name)
    print("--- Tamamlama:", complete_avg)
    print("--- Performans:", performance_avg)

    # Excel'e yazdır
    # (last_row = 1) + (student_i = 0) + 1
    row = last_row + student_i + 1
    ws[f"A{row}"] = date
    ws[f"A{row}"].number_format = "d mmmm yyyy, dddd"
    ws[f"B{row}"] = student_name
    ws[f"C{row}"] = complete_avg
    ws[f"D{row}"] = performance_avg
    # =KÖPRÜ("../images/student-based-reports/Ekran Görüntüsü.png";"Görüntü")
    ws[f"E{row}"] = f'=HYPERLINK(".{img_path}", "Görüntü")'

    # sonraki satırdaki öğrenciye geçmek için
    # bir önceki sayfadaki tabloya geri dönüyoruz
    driver.back()


# tarayıcı kapat
time.sleep(2)
driver.close()

# Excel dosyasını kaydet ve kapat
wb.save(xl_path)
wb.close()
